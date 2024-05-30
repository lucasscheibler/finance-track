from io import BytesIO
from datetime import datetime
from fastapi import UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
from app.database.model.transaction import TransactionModel


class TransactionService():
    '''Transaction Service class'''

    # I need to remove special chars from excel header in order to be albe to import the file.
    # the file needs to be exported from Extratos -> select start date and end date (1year)
    # excel sheet name = Movimentacao

    @staticmethod
    async def import_transactions(file: UploadFile, db: AsyncSession):
        """It read an excel file provided by B3 and import the data into database

        Args:
            file (UploadFile): Excel file provided by B3
            db (AsyncSession): database session
        """        
        buffer = BytesIO(file.file.read())
        wb = load_workbook(buffer, read_only=True)
        ws = wb['Movimentacao']
        file_header = []

        rows = ws.iter_rows(values_only=True)
        file_header = next(rows)
        rows = list(rows)

        for _, row in enumerate(rows):
            if row[file_header.index('Movimentacao')] == 'Transferência - Liquidação':
                transaction = TransactionModel()
                transaction.effective_date = datetime.strptime(row[file_header.index('Data')], '%d/%m/%Y')
                transaction.broker = row[file_header.index('Instituicao')]   
                transaction.code = row[file_header.index('Produto')].split('-')[0].strip()  
                transaction.action = 'BUY' if row[file_header.index('Entrada/Saida')] == "Credito" else 'SELL'
                transaction.num_shares = int(row[file_header.index('Quantidade')])   
                transaction.price = float(row[file_header.index('Preco Unitario')])   
                transaction.total = float(row[file_header.index('Valor da Operacao')])   
                await TransactionModel.save(transaction, db)
                